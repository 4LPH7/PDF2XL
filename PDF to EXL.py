import os
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def extract_cleaned_table_data(pdf_path):
    """
    Extracts and cleans table data from a PDF using pdfplumber.
    Args:
        pdf_path (str): Path to the input PDF file.
    Returns:
        list: A list of tables, where each table is a list of cleaned rows.
    """
    all_tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for table in tables:
                cleaned_table = []
                for row in table:
                    # Clean row: Remove None or empty values and replace with an empty string
                    cleaned_row = [cell if cell is not None and cell.strip() else "" for cell in row]
                    # Add only rows that have meaningful data
                    if any(cleaned_row):  # Ensures at least one non-empty cell exists
                        cleaned_table.append(cleaned_row)
                if cleaned_table:
                    all_tables.append(cleaned_table)
    return all_tables


def write_cleaned_data_to_excel(tables, output_path):
    """
    Writes cleaned data to an Excel file, preserving the table structure and applying formatting.
    Args:
        tables (list): List of tables, each table is a list of rows to be written to Excel.
        output_path (str): Path to save the Excel file.
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Define borders for cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    current_row = 1

    for table in tables:
        # Assuming first row is the header
        headers = table[0]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.border = thin_border

        # Write the rest of the table data
        for row_num, row in enumerate(table[1:], start=current_row + 1):
            for col_num, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

        # Move to the next row after the current table, leaving a blank row
        current_row = row_num + 2

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_path)
    print(f"Data successfully written to '{output_path}'.")


def main():
    pdf_path = "test.pdf"  # Path to the input PDF
    output_path = "output.xlsx"  # Path to the output Excel file

    if not os.path.exists(pdf_path):
        print(f"Error: File '{pdf_path}' not found.")
        return

    print("Extracting and cleaning table data from PDF...")
    tables = extract_cleaned_table_data(pdf_path)

    if not tables:
        print("No valid table data found in the PDF.")
        return

    print("Writing cleaned data to Excel...")
    write_cleaned_data_to_excel(tables, output_path)


if __name__ == "__main__":
    main()